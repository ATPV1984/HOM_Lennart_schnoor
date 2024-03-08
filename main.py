# Python program to read an excel file

# import openpyxl module
# To Install:  openpyxl, Pillow
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.drawing.image import Image
from openpyxl.worksheet import page
from copy import copy
#variables--------------------------------------------------------------------------------------------------------------
#rel_reihe = max_row + 2
tab_start_c = 2
tab_start_r = 2
tab_breite = 4
column_step_lengh = 2

row_average = 0
color_range_for_dbm = ['009600','00FF00','FFFF00','FF9900','FF0000']
color_range_for_evm = ['009600','00FF00','FFFF00','FF9900','FF0000']
dbm_limit_list = [-85,-88,-90,-94]
evm_limit_list = [5,10,15,25]
lac_top_color = ['0287f5','0099b8']
evm_top_color = ['0287f5','0099b8']
id_color = ['8aebe7','42bdb8']
spaltenbreiten = [6,5,15,6,7]

border_styles = ['medium', 'dotted', 'hair', 'mediumDashDotDot',
                 'dashed', 'double', 'thick', 'dashDotDot',
                 'mediumDashDot', 'dashDot', 'mediumDashed',
                 'thin', 'slantDashDot']



#functions------------------------------------------------------
#create a new sheet in main workbook from excel file
def append_excel_file_as_new_sheet(path, file_name_list1, ueberhang_rows, ueberhang_columns):
    print(len(file_name_list1))
    image_counter = 0
    for i in range(0, len(file_name_list1)):
        print(file_name_list1[i])
        #tabelle der TMO wird erstellt
        temp_wb = openpyxl.load_workbook(path + '\\' + file_name_list1[i] + '.xlsx')
        temp_ws = temp_wb.active
        ws_main_temp = wb_main.create_sheet(file_name_list1[i], i*3)
        copy_sheet(temp_ws, ws_main_temp)
        for s in range(1, 6):
            color_id_column(s, 1, ws_main_temp, spaltenbreiten[s-1])
        for c in range(6,ws_main_temp.max_column+1):
            color_lac_column(c, 2, ws_main_temp.max_row+1, ws_main_temp, c%2)
        openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(ws_main_temp, paper_size = 12, orientation='landscape')

        #Bild wird der folgenden Seite beigefügt
        ws_bild = wb_main.create_sheet(file_name_list1[i]+'-Karte', (i*2)+1)
        img = Image(source_path_images + r'\\' + file_name_list[image_counter] + '.jpg')
        image_counter = image_counter + 1
        ws_bild.add_image(img, 'a3')
        openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(ws_bild, paper_size=12,orientation='landscape')
        ws_bild['A1'].value = ueberschrift + file_name_list1[i]
        ws_bild['A1'].font = Font(name='Calibri',
                                 size=11,
                                 bold=True,
                                 italic=False,
                                 vertAlign=None,
                                 underline='none',
                                 strike=False)
        ws_bild.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        #einfügen der Seitenzahl
        ws_main_temp.oddFooter.left.text = '&P/&N'
        ws_main_temp.oddFooter.center.text = '&B' + file_name_list1[i]
        ws_bild.oddFooter.left.text = '&P/&N'
    del wb_main['Sheet']

#Copy a whole page-------------------------------------
def copy_sheet(original_ws, destination_ws):
    max_row = original_ws.max_row +1
    max_column = original_ws.max_column +1
    for y in range(1, 6):
        for r in range(1,max_row):
            destination_ws.cell(row=r, column=y).value = original_ws.cell(row=r, column=y).value

    for y in range(6, max_column,2):
        for r in range(1, max_row):
            t = original_ws.cell(row=r, column=y).value
            if  t != None:
                destination_ws.cell(row=r, column=int(y/2)+3).value = t
            else:
                destination_ws.cell(row=r, column=int(y / 2)+3).fill = PatternFill(patternType='solid',fgColor='333333')

#move a pair of columns within a sheet
def swap_pair_of_columns(top_row, bottom_row, original_column, destinate_column, sheet): #swap eines paares mit einem anderen paar im selben sheet
    temp_col1 = []
    temp_col2 = []
    work_dif = bottom_row - top_row
    for i in range(0, work_dif):
                                    # Save data of Original rows
        temp_col1.append(sheet.cell(row=top_row + i, column=original_column).value)
        temp_col2.append(sheet.cell(row=top_row + i, column=original_column + 1).value)
                                    # original Spalte überschreiben
        sheet.cell(row=top_row + i, column=original_column).value = sheet.cell(row=top_row + i, column=destinate_column).value
        sheet.cell(row=top_row + i, column=original_column + 1).value = sheet.cell(row=top_row + i, column=destinate_column + 1).value
                                    # original Splate an neuer Position wieder herstellen
        sheet.cell(row=top_row + i, column=destinate_column).value = temp_col1[i]
        sheet.cell(row=top_row + i, column=destinate_column + 1).value = temp_col2[i]


#ID-Einfärben
def color_id_column(column, first_row, sheet, breite):
    sheet.column_dimensions[letter_list[column-1]].width = breite
    sheet.cell(row=1,column=column).fill = PatternFill(patternType='solid', fgColor=id_color[first_row%2])
    sheet.cell(row=1, column=column).font = Font(name='Calibri',
                                                 size=11,
                                                 bold=True,
                                                 italic=False,
                                                 vertAlign=None,
                                                 underline='none',
                                                 strike=False,
                                                 color='ffffff')
    sheet.cell(row=1, column=column).border = Border(bottom=Side(border_style='thick'),
                                                     left=Side(border_style='thin', color='000000'),
                                                     right=Side(border_style='thin', color='000000'))
    for i in range(first_row, int(sheet.max_row/2)+1):
        sheet.cell(row=(i*2)+1, column=column).fill = PatternFill(patternType='solid', fgColor=id_color[i%2])
        sheet.cell(row=(i*2), column=column).fill = PatternFill(patternType='solid', fgColor=id_color[i % 2])
        sheet.cell(row=(i*2)+1, column=column).border = Border(bottom=Side(border_style='thin', color='000000'),
                                                                left=Side(border_style='thin', color='000000'),
                                                                right=Side(border_style='thin', color='000000'))
        sheet.cell(row=(i*2), column=column).border = Border(bottom=Side(border_style='thin', color='000000'),
                                                         left=Side(border_style='thin', color='000000'),
                                                         right=Side(border_style='thin', color='000000'),
                                                               top=Side(border_style='thick', color='000000'))

#LAC-Farbgebung für mit Grenzwertüberschreitung
def color_lac_column(column, first_row, last_row, sheet, one_or_zero_for_color):
    sheet.column_dimensions[letter_list[column-1]].width = 14
    sheet.cell(row=1,column=column).fill = PatternFill(patternType='solid', fgColor=lac_top_color[one_or_zero_for_color])
    sheet.cell(row=1, column=column).font = Font(name='Calibri',
                                                 size=11,
                                                 bold=True,
                                                 italic=False,
                                                 vertAlign=None,
                                                 underline='none',
                                                 strike=False,
                                                 color='ffffff')
    sheet.cell(row=1, column=column).border = Border(bottom=Side(border_style='thick'),
                                                     left=Side(border_style='thick', color='000000'),
                                                     right=Side(border_style='thin', color='000000'))
    for i in range(first_row, last_row):
        color = 'ff0000'
        if i%2 == 1:
            sheet.cell(row=i, column=column).border = Border(top=Side(border_style='thin', color='000000'),
                                                             left=Side(border_style='thin', color='000000'),
                                                             right=Side(border_style='thin', color='000000'),
                                                             bottom=Side(border_style='thin', color='000000'))
        else:
            sheet.cell(row=i, column=column).border = Border(top=Side(border_style='thick', color='000000'),
                                                             left=Side(border_style='thin', color='000000'),
                                                             right=Side(border_style='thin', color='000000'),
                                                             bottom=Side(border_style='thin', color='000000'))
        if sheet.cell(row=i,column=column).value != None:
            for y in range(0, 4):
                t = sheet.cell(row=i,column=column).value
                if t != None:
                    if float(t) > dbm_limit_list[y]:
                        color = color_range_for_dbm[y]
                        break
            sheet.cell(row=i, column=column).fill = PatternFill('solid', fgColor=color)

#EVM-Farbgebung für mit Grenzwertüberschreitung
def color_evm_column(column, first_row, last_row, sheet, one_or_zero_for_color):
    sheet.column_dimensions[letter_list[column-1]].width = 10
    sheet.cell(row=1, column=column).fill = PatternFill(patternType='solid', fgColor=lac_top_color[one_or_zero_for_color])
    sheet.cell(row=1, column=column).font = Font(name='Calibri',
                                                 size=11,
                                                 bold=True,
                                                 italic=False,
                                                 vertAlign=None,
                                                 underline='none',
                                                 strike=False,
                                                 color='ffffff')
    sheet.cell(row=1, column=column).border = Border(bottom=Side(border_style='thick'))
    for i in range(first_row, last_row):
        sheet.cell(row=i, column=column).border = Border(bottom=Side(border_style='thin', color='000000'),
                                                         left=Side(border_style='thin', color='000000'),
                                                         right=Side(border_style='thin', color='000000'))
        color = 'ff0000'
        if sheet.cell(row=i,column=column).value != None:
            for y in range(0, 4):
                t = sheet.cell(row=i,column=column).value
                if float(t) < evm_limit_list[y]:
                    color = color_range_for_evm[y]
                    break
            sheet.cell(row=i, column=column).fill = PatternFill('solid', fgColor=color)

#get a list
def get_unique_file_names(directory):
    unique_names = set()

    for filename in os.listdir(directory):
        # Get the base filename without extension
        base_name, extension = os.path.splitext(filename)

        # Check if the base name is not already in the set
        if base_name not in unique_names:
            unique_names.add(base_name)

    return list(sorted(unique_names))



#Setup Preperations-----------------------------------------------------------------------------------------------------
                            # Give the location of the file
source_path_workbooks = r'C:\Users\atpv1\Documents\Arbeit\Lennart\Schnoor\CityCenterBergedorf\Messungen\HOM\\'
source_path_images = r'C:\Users\atpv1\Documents\Arbeit\Lennart\Schnoor\CityCenterBergedorf\Auswertungen\HOM_Bilder_resized\\'
destinate_path = r'C:\Users\atpv1\Documents\Arbeit\Lennart\Schnoor\CityCenterBergedorf\Auswertungen\HOM\\'
destinate_file = r'HOM.xlsx'
                            # To open the workbook
file_name_list = ['HOM']
ueberschrift = 'Anlage 3: Messpunkte '
#get_unique_file_names(source_path_workbooks)#
letter_list = ['a','b','c','d','e','f','g','h','i',
               'j','k','l','m','n','o','p','q','r',
               's','t','u','v','w','x','y','z','aa',
               'ab','ac','ad','ae','af','ag','ah','ai','aj',
               'ak','al','am','an','ao','ap','aq','ar','as','at','au','av',
               'aw','ax','ay','az']
                            # Create the new main Workbook
list_of_column_exceptions = [3,5]
wb_main = Workbook()
ueberhang_rows = 0
ueberhang_columns = 0
append_excel_file_as_new_sheet(source_path_workbooks, file_name_list, ueberhang_rows, ueberhang_columns)



wb_main.save(destinate_file)
